"""Rule Validation tab.

Page:
  GET  /rule-review

API (all read-only against FortiManager; POST is for submitting work items):
  GET  /api/rule-review/adoms
  GET  /api/rule-review/adoms/<adom>/packages
  POST /api/rule-review/parse-import        — parse uploaded CSV or XLSX
  POST /api/rule-review/analyze             — run analysis
  GET  /api/rule-review/zone-status         — is zone policy DB available?
  GET  /api/rule-review/ai-assist-status    — is AI Assist enabled?
  POST /api/rule-review/ai-assist           — single-request AI Assist (planner + LLM narration)
  GET  /api/rule-review/devices             — device:ADOM pairs across accessible ADOMs (AI Assist typeahead)
  POST /api/rule-review/ai-assist-fqdn      — vendor FQDN/wildcard allowlist AI Assist (planner + LLM narration)
"""

import csv
import io
import json as _json

from flask import Blueprint, current_app, jsonify, render_template, request, session

from app import registry
from app.decorators import check_adom_access, tab_required
from app.fmg_client import FMGError
from app.fmg_helpers import make_client
from app.rule_review import analyze_flows, zone_script_available
from app.security import internal_api_error, upstream_api_error

bp = Blueprint("rule_review", __name__)

registry.register("rule_review", "Rule Validation", "rule_review.rule_review_page")


# ── Page ──────────────────────────────────────────────────────────────────────


@bp.route("/rule-review")
@tab_required("rule_review")
def rule_review_page():
    return render_template("rule_review.html", user=session["user"])


# ── API: ADOM list ────────────────────────────────────────────────────────────


@bp.route("/api/rule-review/adoms")
@tab_required("rule_review")
def rr_adoms():
    try:
        from flask import session as _session

        from app.groups import get_allowed_adoms

        allowed = get_allowed_adoms(
            _session.get("user", ""), ad_groups=_session.get("ad_groups", [])
        )
        with make_client() as client:
            raw = client.get_adoms()
        names = sorted(
            a["name"]
            for a in raw
            if isinstance(a, dict)
            and a.get("name")
            and not a["name"].lower().startswith("forti")
        )
        if allowed is not None:
            names = [n for n in names if n in allowed]
        return jsonify(names)
    except FMGError as exc:
        return upstream_api_error("rule_review", exc)
    except Exception as exc:
        return internal_api_error("rule_review", exc)


# ── API: device list across accessible ADOMs (AI Assist typeahead) ────────────


@bp.route("/api/rule-review/devices")
@tab_required("rule_review")
def rr_devices():
    """List every device the current user can target, paired with its ADOM.

    Backs the AI Assist "Target firewall(s)" typeahead — the field expects
    DEVICE:ADOM, and this endpoint saves the engineer a trip to the
    Firewalls tab to look up which ADOM a device lives in.
    """
    try:
        from flask import session as _session

        from app.groups import get_allowed_adoms

        allowed = get_allowed_adoms(
            _session.get("user", ""), ad_groups=_session.get("ad_groups", [])
        )
        with make_client() as client:
            raw_adoms = client.get_adoms()
            adom_names = sorted(
                a["name"]
                for a in raw_adoms
                if isinstance(a, dict)
                and a.get("name")
                and not a["name"].lower().startswith("forti")
            )
            if allowed is not None:
                adom_names = [n for n in adom_names if n in allowed]

            results = []
            for adom in adom_names:
                for d in client.get_devices(adom):
                    if isinstance(d, dict) and d.get("name"):
                        results.append({"device": d["name"], "adom": adom})
        results.sort(key=lambda r: (r["device"], r["adom"]))
        return jsonify(results)
    except FMGError as exc:
        return upstream_api_error("rule_review", exc)
    except Exception as exc:
        return internal_api_error("rule_review", exc)


# ── API: package list ─────────────────────────────────────────────────────────


@bp.route("/api/rule-review/adoms/<adom>/packages")
@tab_required("rule_review")
def rr_packages(adom: str):
    if err := check_adom_access(adom):
        return err
    try:
        with make_client() as client:
            raw = client.get_policy_packages(adom)
        packages = [
            {"name": p["name"], "path": p.get("path", p["name"])}
            for p in raw
            if isinstance(p, dict)
            and p.get("name")
            and (p.get("type") or "").lower() != "folder"
        ]
        return jsonify(packages)
    except FMGError as exc:
        return upstream_api_error("rule_review", exc)
    except Exception as exc:
        return internal_api_error("rule_review", exc)


# ── API: parse import file ────────────────────────────────────────────────────

# Shared by rr_parse_import and rr_ai_assist_fqdn — both accept .xlsx uploads
# and must apply the same content-type and size validation before openpyxl.
_ALLOWED_XLSX_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/octet-stream",
}


def _stream_size(stream) -> int:
    """Byte length of an uploaded file stream, leaving the position unchanged."""
    pos = stream.tell()
    stream.seek(0, 2)
    size = stream.tell()
    stream.seek(pos)
    return size


@bp.route("/api/rule-review/parse-import", methods=["POST"])
@tab_required("rule_review")
def rr_parse_import():
    """Accept a CSV or XLSX upload, return parsed rows.

    Expected columns (case-insensitive, flexible order):
      src / source / source_ip
      dst / destination / dest / destination_ip / dest_ip
      port / service / ports / services
      comment / comments / note / notes  (optional)
    """
    if "file" not in request.files:
        return jsonify({"error": "No file uploaded"}), 400

    f = request.files["file"]
    filename = (f.filename or "").lower()
    mimetype = (f.mimetype or "").lower()

    allowed_csv_types = {
        "text/csv",
        "application/csv",
        "application/vnd.ms-excel",
        "application/octet-stream",
    }
    allowed_xlsx_types = _ALLOWED_XLSX_TYPES

    max_bytes = int(current_app.config.get("MAX_CONTENT_LENGTH", 4 * 1024 * 1024))
    if _stream_size(f.stream) > max_bytes:
        return jsonify({"error": "Uploaded file is too large"}), 413

    rows: list[dict] = []
    errors: list[str] = []

    if filename.endswith(".csv"):
        if mimetype not in allowed_csv_types:
            return jsonify({"error": "Unsupported CSV content type"}), 400
        try:
            content = f.read().decode("utf-8-sig", errors="replace")
            reader = csv.DictReader(io.StringIO(content))
            rows, errors = _parse_rows(reader)
        except Exception:
            return jsonify({"error": "CSV parse error"}), 400

    elif filename.endswith((".xls", ".xlsx")):
        if mimetype not in allowed_xlsx_types:
            return jsonify({"error": "Unsupported XLSX content type"}), 400
        try:
            import openpyxl

            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
            ws = wb.active
            headers = [
                str(c.value or "").strip().lower()
                for c in next(ws.iter_rows(min_row=1, max_row=1))
            ]
            dict_rows = []
            for row in ws.iter_rows(min_row=2, values_only=True):
                dict_rows.append(
                    {
                        headers[i]: (str(v).strip() if v is not None else "")
                        for i, v in enumerate(row)
                    }
                )
            rows, errors = _parse_rows(dict_rows)
        except ImportError:
            return jsonify(
                {"error": "openpyxl not installed — XLSX import unavailable"}
            ), 500
        except Exception:
            return jsonify({"error": "XLSX parse error"}), 400
    else:
        return jsonify(
            {"error": "Unsupported file type. Upload a .csv or .xlsx file."}
        ), 400

    return jsonify({"rows": rows, "errors": errors})


_SRC_ALIASES = {"src", "source", "source_ip", "src_ip"}
_DST_ALIASES = {"dst", "destination", "dest", "destination_ip", "dest_ip", "dst_ip"}
_SVC_ALIASES = {"port", "ports", "service", "services", "svc"}
_CMT_ALIASES = {"comment", "comments", "note", "notes"}


def _canonical_header(raw: str) -> str:
    key = raw.strip().lower().replace(" ", "_").replace("-", "_")
    if key in _SRC_ALIASES:
        return "src"
    if key in _DST_ALIASES:
        return "dst"
    if key in _SVC_ALIASES:
        return "service"
    if key in _CMT_ALIASES:
        return "comment"
    return key


def _parse_rows(reader) -> tuple[list[dict], list[str]]:
    rows: list[dict] = []
    errors: list[str] = []
    for line_no, raw_row in enumerate(reader, start=2):
        row = {
            _canonical_header(k): (v or "").strip()
            for k, v in (raw_row.items() if hasattr(raw_row, "items") else {}.items())
        }
        src = row.get("src", "")
        dst = row.get("dst", "")
        svc = row.get("service", "")
        if not src and not dst:
            continue  # skip blank rows
        if not src:
            errors.append(f"Row {line_no}: missing source IP")
            continue
        if not dst:
            errors.append(f"Row {line_no}: missing destination IP")
            continue
        rows.append(
            {"src": src, "dst": dst, "service": svc, "comment": row.get("comment", "")}
        )
    return rows, errors


# ── API: zone policy status ───────────────────────────────────────────────────


@bp.route("/api/rule-review/zone-status")
@tab_required("rule_review")
def rr_zone_status():
    return jsonify({"available": zone_script_available()})


# ── API: analyze ─────────────────────────────────────────────────────────────


@bp.route("/api/rule-review/analyze", methods=["POST"])
@tab_required("rule_review")
def rr_analyze():
    """Run the policy review analysis.

    Request body::
        {
            "flows": [{"src": "...", "dst": "...", "service": "...", "comment": "..."}, ...],
            "packages": [{"adom": "...", "name": "...", "path": "...", "device": "..."}, ...]
        }
    """
    data = request.get_json(silent=True) or {}
    flows = data.get("flows", [])
    packages = data.get("packages", [])

    if not flows:
        return jsonify({"error": "No flows provided"}), 400
    if not packages:
        return jsonify({"error": "No policy packages selected"}), 400

    # Collect unique ADOMs to minimise API calls
    adoms = list(dict.fromkeys(p["adom"] for p in packages if p.get("adom")))

    # Enforce ADOM access for every ADOM referenced
    for adom in adoms:
        if err := check_adom_access(adom):
            return err

    try:
        with make_client() as client:
            # Fetch policies for each package
            policies_by_pkg: dict[str, list] = {}
            for pkg in packages:
                adom = pkg["adom"]
                path = pkg["path"]
                key = f"{adom}/{path}"
                try:
                    policies_by_pkg[key] = client.get_policies(adom, path)
                except Exception:
                    policies_by_pkg[key] = []

            # Fetch address and service objects per ADOM
            addr_objects: list = []
            addr_groups: list = []
            svc_objects: list = []
            svc_groups: list = []
            seen_adoms: set[str] = set()
            for adom in adoms:
                if adom in seen_adoms:
                    continue
                seen_adoms.add(adom)
                addr_objects.extend(client.get_address_objects(adom))
                addr_groups.extend(client.get_address_groups(adom))
                svc_objects.extend(client.get_service_objects(adom))
                svc_groups.extend(client.get_service_groups(adom))

            # Fetch routing + interface data for path-relevance check.
            # Resolve devices from package scope members; fall back to pkg["device"] if set.
            routing_by_device: dict[str, dict] = {}
            for pkg in packages:
                adom = pkg["adom"]
                path = pkg["path"]
                device = pkg.get("device", "")

                # Try to enumerate scope members of the package first
                scope = client.get_pkg_scope_members(adom, path)
                device_names = (
                    [m.get("name", m) if isinstance(m, dict) else str(m) for m in scope]
                    if scope
                    else []
                )

                if not device_names and device:
                    device_names = [device]

                for dev_name in device_names:
                    if dev_name in routing_by_device:
                        continue
                    try:
                        ifaces = client.get_device_interfaces_all_vdoms(adom, dev_name)
                        routes = client.get_device_routes_all_vdoms(adom, dev_name)
                        routing_by_device[dev_name] = {
                            "interfaces": ifaces,
                            "routes": routes,
                        }
                        # Back-fill pkg["device"] so the engine can look it up
                        if not pkg.get("device"):
                            pkg["device"] = dev_name
                    except Exception:
                        routing_by_device[dev_name] = {"interfaces": [], "routes": []}

    except FMGError as exc:
        return upstream_api_error("rule_review", exc)
    except Exception as exc:
        return internal_api_error("rule_review", exc)

    results = analyze_flows(
        requested_flows=flows,
        packages=packages,
        policies_by_pkg=policies_by_pkg,
        addr_objects=addr_objects,
        addr_groups=addr_groups,
        svc_objects=svc_objects,
        svc_groups=svc_groups,
        routing_by_device=routing_by_device,
    )

    return jsonify({"results": results, "zone_available": zone_script_available()})


# ── AI Assist ─────────────────────────────────────────────────────────────────


@bp.route("/api/rule-review/ai-assist-status")
@tab_required("rule_review")
def rr_ai_assist_status():
    from app.app_settings import get_setting

    return jsonify({"available": get_setting("ai_assist_enabled", False)})


@bp.route("/api/rule-review/ai-assist", methods=["POST"])
@tab_required("rule_review")
def rr_ai_assist():
    """AI Assist: run plan_change deterministically, then narrate the result
    with the configured LLM. The deterministic result is always returned;
    narration is best-effort and degrades gracefully on failure — the LLM
    never computes or edits any value in the plan, it only explains it."""
    from app.app_settings import get_setting

    if not get_setting("ai_assist_enabled", False):
        return jsonify({"error": "AI Assist is not enabled"}), 503

    data = request.get_json(silent=True) or {}
    src = data.get("src", "")
    dst = data.get("dst", "")
    service = data.get("service", "")
    firewalls_raw = data.get("firewalls", [])
    ticket_id = data.get("ticket_id", "")
    justification = data.get("justification", "")
    src_group = data.get("src_group", "")
    dst_group = data.get("dst_group", "")

    if not src or not dst or not service or not firewalls_raw:
        return jsonify({"error": "src, dst, service, and firewalls are required"}), 400

    for fw in firewalls_raw:
        if not fw.get("device") or not fw.get("adom"):
            return jsonify(
                {
                    "error": "Each target firewall must include both a device and an ADOM "
                    "(format: DEVICE:ADOM) — got an entry missing one or the other."
                }
            ), 400
        if err := check_adom_access(fw["adom"]):
            return err

    from app.planner.engine import plan_change
    from app.planner.models import PlannerDataError, TargetFirewall

    targets = [
        TargetFirewall(device=fw["device"], adom=fw["adom"]) for fw in firewalls_raw
    ]

    path_relevance: dict = {}
    try:
        with make_client() as fmg:
            plan = plan_change(
                src=src,
                dst=dst,
                service=service,
                firewalls=targets,
                justification=justification,
                ticket_id=ticket_id,
                src_group=src_group,
                dst_group=dst_group,
                fmg_client=fmg,
            )

            # Path-relevance ("is this firewall actually in the traffic path")
            # has no equivalent in the ported planner — it's 4THealth+-specific
            # and wraps the planner's output, same as it already wraps the
            # existing bulk-analysis engine. Scoped to the single-src/single-dst
            # case (the common one); multi-value requests skip this check
            # rather than guessing which pair to report on.
            #
            # This entire block is best-effort and MUST NEVER raise: the plan
            # computed above has already succeeded, and losing it because an
            # advisory annotation failed would violate the core guarantee that
            # the deterministic result always renders. Every failure mode here
            # — a bad FMG call, a malformed interface/route shape,  anything
            # unanticipated — degrades to a per-device note, never a 500.
            srcs_list = [s.strip() for s in src.split(",") if s.strip()]
            dsts_list = [d.strip() for d in dst.split(",") if d.strip()]
            if len(srcs_list) == 1 and len(dsts_list) == 1:
                from app.rule_review import check_path_relevance

                for target in targets:
                    try:
                        interfaces = fmg.get_device_interfaces_all_vdoms(
                            target.adom, target.device
                        )
                        routes = fmg.get_device_routes_all_vdoms(
                            target.adom, target.device
                        )
                        path_relevance[target.device] = check_path_relevance(
                            srcs_list[0],
                            dsts_list[0],
                            interfaces,
                            routes,
                        )
                    except Exception:
                        path_relevance[target.device] = {
                            "in_path": None,
                            "confidence": "low",
                            "notes": ["Could not determine path relevance"],
                        }
    except PlannerDataError as exc:
        return jsonify({"error": str(exc), "source": exc.source}), 502
    except FMGError as exc:
        return upstream_api_error("rule_review", exc)
    except Exception as exc:
        return internal_api_error("rule_review", exc)

    plan_dict = plan.to_dict()

    narrative = None
    narrative_error = None
    try:
        from app.llm import get_provider

        provider = get_provider()
        narrative = provider.narrate(
            system_prompt=(
                "You are a firewall change analyst assistant. You are given a "
                "structured, already-computed change plan as JSON. Write a clear, "
                "concise report for a peer reviewer: summarize the verdict, the "
                "required change per firewall, risk level, and approval "
                "requirements. Never invent or change any value in the plan — "
                "only explain it in prose."
            ),
            user_prompt=_json.dumps(plan_dict, default=str),
            feature="rule_review_ai_assist",
            user=session.get("user"),
        )
    except Exception as exc:
        narrative_error = str(exc)

    return jsonify(
        {
            "plan": plan_dict,
            "narrative": narrative,
            "narrative_error": narrative_error,
            "path_relevance": path_relevance,
        }
    )


# ── AI Assist: FQDN allowlist ───────────────────────────────────────────────


@bp.route("/api/rule-review/ai-assist-fqdn", methods=["POST"])
@tab_required("rule_review")
def rr_ai_assist_fqdn():
    """AI Assist (FQDN allowlist mode): run plan_fqdn_change deterministically,
    then narrate the result with the configured LLM. Same guarantees as
    rr_ai_assist — the deterministic plan always returns; narration is
    best-effort."""
    from app.app_settings import get_setting

    if not get_setting("ai_assist_enabled", False):
        return jsonify({"error": "AI Assist is not enabled"}), 503

    from app.planner.fqdn_intake import parse_fqdn_rows, parse_fqdn_xlsx
    from app.planner.models import FQDNAllowlistRequest

    is_multipart = "file" in request.files

    def _validate_firewalls(fws):
        """Type- and presence-check every firewall entry before it is used to
        build the FQDNAllowlistRequest or the ADOM-access loop — a firewall
        entry missing "device" or "adom" (or a non-list `firewalls` value)
        must never reach a dict-key lookup that would raise an unhandled
        TypeError/KeyError (and surface as a raw 500)."""
        if not isinstance(fws, list):
            return jsonify({"error": "firewalls must be a JSON array"}), 400
        for fw in fws:
            if not isinstance(fw, dict) or not fw.get("device") or not fw.get("adom"):
                return jsonify(
                    {
                        "error": "Each target firewall must include both a device and an ADOM "
                        "(format: DEVICE:ADOM) — got an entry missing one or the other."
                    }
                ), 400
        return None

    if is_multipart:
        src_ip = request.form.get("src_ip", "")
        ticket_id = request.form.get("ticket_id", "")
        try:
            firewalls_raw = _json.loads(request.form.get("firewalls", "[]"))
        except ValueError:
            return jsonify({"error": "firewalls must be a JSON array"}), 400
        # Same required-field contract as the JSON branch below.
        if not src_ip or not firewalls_raw:
            return jsonify(
                {"error": "src_ip and at least one target firewall are required"}
            ), 400
        if err := _validate_firewalls(firewalls_raw):
            return err

        # Same upload validation as rr_parse_import — extension + content type
        # are checked before the file ever reaches openpyxl.
        upload = request.files["file"]
        if not (upload.filename or "").lower().endswith(".xlsx"):
            return jsonify({"error": "Only .xlsx uploads are supported"}), 400
        if (upload.mimetype or "").lower() not in _ALLOWED_XLSX_TYPES:
            return jsonify({"error": "Unsupported XLSX content type"}), 400
        if _stream_size(upload.stream) > int(
            current_app.config.get("MAX_CONTENT_LENGTH", 4 * 1024 * 1024)
        ):
            return jsonify({"error": "Uploaded file is too large"}), 413

        try:
            parsed = parse_fqdn_xlsx(
                upload,
                src_ip=src_ip,
                ticket_id=ticket_id,
                firewalls=[f"{fw['device']}:{fw['adom']}" for fw in firewalls_raw],
            )
        except Exception as exc:
            return jsonify({"error": f"Could not parse uploaded .xlsx: {exc}"}), 400
        if not parsed.entries:
            return jsonify(
                {
                    "error": "No valid FQDN rows found in the uploaded file",
                    "warnings": parsed.warnings,
                }
            ), 400
        vendor = request.form.get("vendor", "") or parsed.vendor
        category = request.form.get("category", "") or parsed.category
    else:
        data = request.get_json(silent=True) or {}
        vendor = data.get("vendor", "")
        category = data.get("category", "")
        src_ip = data.get("src_ip", "")
        ticket_id = data.get("ticket_id", "")
        firewalls_raw = data.get("firewalls", [])
        entries_raw = data.get("entries", [])

        if not src_ip or not firewalls_raw or not entries_raw:
            return jsonify(
                {"error": "src_ip, firewalls, and at least one entry are required"}
            ), 400
        if not isinstance(entries_raw, list):
            return jsonify({"error": "entries must be a JSON array"}), 400

        if err := _validate_firewalls(firewalls_raw):
            return err

        # Route the manual-entry path through the same parser as the .xlsx
        # upload path (design decision 2: "Both paths resolve to the same
        # FQDNAllowlistRequest/FQDNEntry structures, built by one new
        # app/planner/fqdn_intake.py module"). The parser owns illegal-character
        # rejection, port coercion, and protocol validation, so a hostile or
        # malformed payload produces a 400 with warnings rather than an
        # uncaught ValueError → 500.
        rows: list[dict] = []
        for e in entries_raw:
            if not isinstance(e, dict):
                continue
            ports_val = e.get("ports", "")
            if isinstance(ports_val, (list, tuple)):
                ports_val = ",".join(str(p) for p in ports_val)
            rows.append(
                {
                    "fqdn": e.get("fqdn", ""),
                    "ports": str(ports_val),
                    "protocol": e.get("protocol", "TCP"),
                    "required": e.get("required", True),
                    "comment": e.get("comment", ""),
                    # vendor/category are request-level, not per-entry, in the
                    # JSON payload — the parser reads them off the rows.
                    "vendor": vendor,
                    "category": category,
                }
            )

        parsed = parse_fqdn_rows(
            rows,
            src_ip=src_ip,
            ticket_id=ticket_id,
            firewalls=[f"{fw['device']}:{fw['adom']}" for fw in firewalls_raw],
        )
        if not parsed.entries:
            return jsonify(
                {
                    "error": "No valid FQDN entries provided",
                    "warnings": parsed.warnings,
                }
            ), 400

    fqdn_request = FQDNAllowlistRequest(
        vendor=vendor,
        category=category,
        src_ip=src_ip,
        ticket_id=ticket_id,
        firewalls=[f"{fw['device']}:{fw['adom']}" for fw in firewalls_raw],
        entries=parsed.entries,
    )
    intake_warnings = list(parsed.warnings)
    intake_missing_fields = list(parsed.missing_fields)

    for fw in firewalls_raw:
        if err := check_adom_access(fw["adom"]):
            return err

    from app.planner.engine import plan_fqdn_change, to_fqdn_report_payload
    from app.planner.models import PlannerDataError

    try:
        with make_client() as fmg:
            plan = plan_fqdn_change(fqdn_request, fmg_client=fmg)
    except PlannerDataError as exc:
        return jsonify({"error": str(exc), "source": exc.source}), 502
    except FMGError as exc:
        return upstream_api_error("rule_review", exc)
    except Exception as exc:
        return internal_api_error("rule_review", exc)

    plan_dict = to_fqdn_report_payload(plan)
    plan_dict["intake_warnings"] = intake_warnings
    plan_dict["intake_missing_fields"] = intake_missing_fields

    narrative = None
    narrative_error = None
    try:
        from app.llm import get_provider

        provider = get_provider()
        narrative = provider.narrate(
            system_prompt=(
                "You are a firewall change analyst assistant. You are given a "
                "structured, already-computed FQDN/wildcard allowlist change plan "
                "as JSON — one entry per target firewall with coverage status and "
                "any proposed address objects/group/policy. Write a clear, concise "
                "report for a peer reviewer: summarize coverage per firewall, what "
                "needs to be created, and any warnings. Never invent or change any "
                "value in the plan — only explain it in prose."
            ),
            user_prompt=_json.dumps(plan_dict, default=str),
            feature="rule_review_ai_assist_fqdn",
            user=session.get("user"),
        )
    except Exception as exc:
        narrative_error = str(exc)

    return jsonify(
        {
            "plan": plan_dict,
            "narrative": narrative,
            "narrative_error": narrative_error,
        }
    )
